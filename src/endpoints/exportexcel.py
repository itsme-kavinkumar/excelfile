from fastapi import APIRouter,HTTPException,Depends
import openpyxl.workbook
from src.models.mysql.exportexcel import*
from database import get_db
from typing import AsyncGenerator
import openpyxl
from openpyxl.styles import alignment,PatternFill,Font,borders,Side
import traceback
import uuid,os
router=APIRouter()
# from test import currency
@router.post('/export-data')
async def ExportToExcel(db:AsyncGenerator=Depends(get_db)):
    try:
        result=await GetQuery(db)

        main_file_path = r"Book1.xlsx"        
        workbook = openpyxl.load_workbook(main_file_path)
        main_file = workbook.active
        row=4
        for data in result:
            main_file.cell(row,column=1).value=data['studentCode']
            main_file.cell(row,column=2).value=data['Name']
            main_file.cell(row,column=3).value=data['sem1']
            main_file.cell(row,column=4).value=data['sem2']
            main_file.cell(row,column=5).value=data['sem3']
            main_file.cell(row,column=6).value=data['sem4']
            main_file.cell(row,column=7).value=data['sem5']
            main_file.cell(row,column=8).value=data['sem6']
            main_file.cell(row,column=9).value=data['total_mark']
            row+=1
      
        workbook.save(generate_filename("student_mark_list", "xlsx"))
        print(f"sheet downloaded sucessfully")
        return {'msg':"sheet downloaded sucessfully"}
    except Exception as e :
        print("+++++++++++",traceback.extract_tb(e.__traceback__))
        return HTTPException(status_code=500,detail=str(e))
    


def generate_filename(name,type):
    filepath= f"{name}"+str(uuid.uuid1())+'.'+ type
    while os.path.exists(filepath):
        filepath= "student_mark_list"+str(uuid.uuid1())+'.'+ type
    return filepath  
   

def CreateExcel():
    try:
        workbook= openpyxl.Workbook()
        sheet= workbook.active
        mrg=sheet.merge_cells( start_row=1,start_column=1,end_row=2,end_column=9)
        header=sheet.cell(row=1,column=1)
        header.alignment=alignment.Alignment(horizontal="center")
        header.fill=PatternFill(start_color="0C920F",fill_type='solid')
        header.font=Font(size='20',bold=True,color="FFFFFF",name="Arial")

        for rows in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=9):
            for cell in rows:
                cell.border=borders.Border(left=Side(border_style=borders.BORDER_THICK,color='EB430F'),
                                    right=Side(border_style=borders.BORDER_THICK,color='EB430F'),
                                    top=Side(border_style=borders.BORDER_THICK,color='EB430F'),
                                    bottom=Side(border_style=borders.BORDER_THICK,color='EB430F'))
    
        header.value="Students Mark List"
        # header.value=currency(450000)
        sheet.cell(row=4,column=1).value="Student Code"
        sheet.cell(row=4,column=2).value="Student Name"
        sheet.cell(row=4,column=3).value="Sem 1"
        sheet.cell(row=4,column=4).value="Sem 2"
        sheet.cell(row=4,column=5).value="Sem 3"
        sheet.cell(row=4,column=6).value="Sem 4"
        sheet.cell(row=4,column=7).value="Sem 5"
        sheet.cell(row=4,column=8).value="Sem 6"
        sheet.cell(row=4,column=9).value="Total Marks"
        
        adjust_columns=['A','B','C','D','E','F','G','H','I']

        for column in adjust_columns:
            max_length = 0
            for cells in sheet[column]:
                if cells.value is not None: 
                    cell_length= len(str(cells.value))
                    if cell_length >max_length:
                        max_length+=cell_length
            sheet.column_dimensions[column].width=max_length+2

        for rows in sheet.iter_rows(min_row=4,max_row=4,min_col=1,max_col=9):
            for cells in rows:
                cells.fill=PatternFill(start_color="8383E1",fill_type="solid")
                cells.font=Font(bold=True)
        
        filename=generate_filename('student-template', 'xlsx')
        workbook.save(filename)
        workbook.close()
        
        print("===sucess======")
        return filename
    except Exception as e:
        print('---------',str(e),traceback.extract_tb(e.__traceback__))
        return str(e)






